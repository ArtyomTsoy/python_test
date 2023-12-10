import openpyxl

file_names = ['1111.xlsx', '2222.xlsx', '3333.xlsx']
data = []

for file_name in file_names:
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    sheet_data = []
    for row in sheet.iter_rows(values_only=True):
        sheet_data.append(row)
    data.extend(sheet_data)

data.sort(reverse=True)

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

for row_idx, row_data in enumerate(data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = new_sheet.cell(row=row_idx, column=col_idx)
        cell.value = value

        cell.font = openpyxl.styles.Font(bold=True)
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )

new_workbook.save('combined_sorted.xlsx')
